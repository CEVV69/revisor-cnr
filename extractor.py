"""Extracción de texto desde PDF, Word, Excel y ZIP"""
from pathlib import Path
import re

# ─── Mapa de anexos SEP → tipo de documento ───────────────────────────────────
ANEXOS_SEP = {
    "9.1":    ("plano_ubicacion",       "Anexo 9.1 - Plano de ubicación del proyecto"),
    "9.2":    ("identificacion_riego",  "Anexo 9.2 - Identificación del área de riego"),
    "9.4.2":  ("pruebas_bombeo",        "Anexo 9.4.2 - Prueba de bombeo"),
    "9.4":    ("estudio_hidrologico",   "Anexo 9.4 - Análisis Hidrológico"),
    "9.5":    ("diseno_hidraulico",     "Anexo 9.5 - Diseño y cálculos hidráulicos"),
    "9.6":    ("estudios_complementarios", "Anexo 9.6 - Estudios y diseño complementarios"),
    "9.8.1":  ("cronograma",            "Anexo 9.8.1 - Cronograma"),
    "9.8":    ("especificaciones_tecnicas", "Anexo 9.8 - Especificaciones técnicas"),
    "9.9":    ("cubicaciones",          "Anexo 9.9 - Cubicaciones"),
    "9.10.1": ("presupuesto",           "Anexo 9.10.1 - Presupuesto detallado de obras"),
    "9.10.2": ("presupuesto_electrico", "Anexo 9.10.2 - Presupuesto detallado electrificación"),
    "9.10.3": ("cotizaciones_facturas", "Anexo 9.10.3 - Cotizaciones y Facturas"),
    "9.10.4": ("cotizaciones",          "Anexo 9.10.4 - Cotizaciones"),
    "9.10.5": ("declaracion_iva",       "Anexo 9.10.5 - Declaración No Contribuyente IVA"),
    "9.12.1.2": ("planos_obras_civiles","Anexo 9.12.1.2 - Planos Obras Civiles"),
    "9.12.1":   ("planos_tecnificacion","Anexo 9.12.1.1 - Planos proyecto tecnificación"),
    "9.12":     ("planos_tecnificacion","Anexo 9.12 - Planos tecnificación"),
    "9.13.1": ("memoria_superficies",   "Anexo 9.13.1 - Memoria de cálculo de superficies"),
    "9.13.2": ("estudio_suelos",        "Anexo 9.13.2 - Estudio de suelos"),
    "9.14":   ("evaluacion_social",     "Anexo 9.14 - Evaluación Social MIDESO"),
}

# Orden de búsqueda: más específico primero
ANEXOS_ORDEN = sorted(ANEXOS_SEP.keys(), key=lambda x: len(x), reverse=True)


def detectar_anexo(nombre_archivo: str) -> tuple:
    """
    Detecta el tipo de anexo desde el nombre del archivo.
    Intento 1: patrón completo "9.X.Y" (ej: 9.10.1).
    Intento 2: patrón sin "9." inicial para claves multi-nivel (ej: 10.1, 8.1, 12.1.2).
    Retorna (tipo_doc, label) o ("otro", "Documento sin clasificar").
    """
    nombre = nombre_archivo.lower()

    # Intento 1: patrón completo con "9."
    for clave in ANEXOS_ORDEN:
        patron = clave.replace(".", r"[\.\-_\s]")
        if re.search(rf"9{patron[1:]}", nombre):
            return ANEXOS_SEP[clave]

    # Intento 2: sin el "9." inicial — solo para claves multi-nivel (ej: 9.10.1 → 10.1)
    for clave in ANEXOS_ORDEN:
        sufijo = clave[2:]          # quita "9."
        if "." not in sufijo:       # evitar "1","4","5"… demasiado ambiguos
            continue
        patron = sufijo.replace(".", r"[\.\-_\s]")
        if re.search(rf"(?<!\d){patron}(?!\d)", nombre):
            return ANEXOS_SEP[clave]

    return ("otro", "Documento sin clasificar")


# Tope de texto extraído que se GUARDA por documento. Antes era 5.000 caracteres (~2 páginas),
# lo que capaba silenciosamente todo lo que la IA podía ver de cada documento — el análisis por
# ítem reparte hasta 45.000 caracteres entre los documentos del grupo y las extracciones
# numéricas buscan datos que suelen estar al final (resultados/conclusiones). 60.000 caracteres
# cubre el documento típico completo; si es más largo, se conserva inicio (75%) + final (25%),
# igual que _truncar_inteligente en analyzer.py.
MAX_CHARS_GUARDADO = 60000


def truncar_texto_guardado(texto: str) -> str:
    """Trunca el texto extraído antes de guardarlo, conservando inicio y final del documento
    (las conclusiones/resultados suelen ir al final). No toca marcadores especiales."""
    if not texto or len(texto) <= MAX_CHARS_GUARDADO or texto == "__PDF_ESCANEADO__":
        return texto
    inicio = int(MAX_CHARS_GUARDADO * 0.75)
    fin = MAX_CHARS_GUARDADO - inicio
    omitidos = len(texto) - MAX_CHARS_GUARDADO
    return (texto[:inicio]
            + f"\n\n[... {omitidos:,} caracteres omitidos — documento muy largo ...]\n\n"
            + texto[-fin:])


def extract_text(filepath: str, ext: str) -> str:
    ext = ext.lower()
    try:
        if ext == ".pdf":
            return _from_pdf(filepath)
        elif ext in (".doc", ".docx"):
            return _from_word(filepath)
        elif ext in (".xls", ".xlsx"):
            return _from_excel(filepath)
        else:
            return f"[Formato {ext} no soportado para extracción de texto automática]"
    except Exception as e:
        return f"[Error al extraer texto: {str(e)}]"


def extract_zip(zip_path: str, dest_dir: str) -> list:
    """
    Extrae un ZIP y retorna lista de dicts con info de cada archivo extraído.
    Ignora archivos ocultos, directorios y formatos no soportados.
    """
    import zipfile
    dest = Path(dest_dir)
    dest.mkdir(parents=True, exist_ok=True)

    FORMATOS_SOPORTADOS = {".pdf", ".doc", ".docx", ".xls", ".xlsx"}
    archivos = []

    with zipfile.ZipFile(zip_path, "r") as zf:
        for member in zf.infolist():
            nombre = Path(member.filename).name
            # Ignorar carpetas, archivos ocultos y formatos no soportados
            if member.is_dir():
                continue
            if nombre.startswith(".") or nombre.startswith("__"):
                continue
            ext = Path(nombre).suffix.lower()
            if ext not in FORMATOS_SOPORTADOS:
                continue

            # Extraer archivo plano (sin subcarpetas)
            destino = dest / nombre
            # Si hay duplicado, añadir sufijo
            contador = 1
            while destino.exists():
                destino = dest / f"{Path(nombre).stem}_{contador}{ext}"
                contador += 1

            with zf.open(member) as src, open(destino, "wb") as dst:
                dst.write(src.read())

            tipo_doc, label = detectar_anexo(nombre)
            texto = extract_text(str(destino), ext)

            archivos.append({
                "nombre_original": nombre,
                "filename": destino.name,
                "tipo_doc": tipo_doc,
                "label": label,
                "texto_extraido": truncar_texto_guardado(texto),
            })

    return archivos


def _from_pdf(filepath: str) -> str:
    """Extrae texto de PDF. Si el PDF es escaneado (imagen), retorna marcador especial."""
    import fitz  # pymupdf
    doc = fitz.open(filepath)
    text_pages = []
    for page in doc:
        t = page.get_text().strip()
        if t:
            text_pages.append(t)
    doc.close()

    if text_pages:
        return "\n".join(text_pages)

    # PDF escaneado — sin texto digital
    return "__PDF_ESCANEADO__"


def render_pdf_as_images(filepath: str, max_pages: int = 4, zoom: float = 0.8) -> list:
    """
    Renderiza páginas de un PDF como imágenes base64 para visión de Claude.
    zoom=0.8 → ~58 dpi, suficiente para lectura de texto y tablas, menor costo en tokens.
    Límite API Claude: 5 MB por imagen.
    """
    import fitz
    import base64
    MAX_BYTES_POR_IMAGEN = 4 * 1024 * 1024  # 4 MB de margen
    doc = fitz.open(filepath)
    images = []
    for i, page in enumerate(doc):
        if i >= max_pages:
            break
        mat = fitz.Matrix(zoom, zoom)
        pix = page.get_pixmap(matrix=mat)
        img_bytes = pix.tobytes("jpeg", jpg_quality=70)
        if len(img_bytes) > MAX_BYTES_POR_IMAGEN:
            mat2 = fitz.Matrix(0.6, 0.6)
            pix2 = page.get_pixmap(matrix=mat2)
            img_bytes = pix2.tobytes("jpeg", jpg_quality=65)
        images.append(base64.standard_b64encode(img_bytes).decode("utf-8"))
    doc.close()
    return images


def _from_word(filepath: str) -> str:
    from docx import Document
    doc = Document(filepath)
    return "\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _from_excel(filepath: str) -> str:
    ext = Path(filepath).suffix.lower()
    if ext == ".xls":
        return _from_xls(filepath)
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    lines = []
    for sheet in wb.worksheets:
        lines.append(f"=== Hoja: {sheet.title} ===")
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(c) if c is not None else "" for c in row)
            if row_text.strip(" |"):
                lines.append(row_text)
    return "\n".join(lines)


def _from_xls(filepath: str) -> str:
    import xlrd
    wb = xlrd.open_workbook(filepath)
    lines = []
    for sheet in wb.sheets():
        lines.append(f"=== Hoja: {sheet.name} ===")
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            row_text = " | ".join(str(c) if c != "" else "" for c in row)
            if row_text.strip(" |"):
                lines.append(row_text)
    return "\n".join(lines)


# ─── Tabla de precios referenciales (categoria, item, unidad, precio) ─────────
# A diferencia de _from_excel (texto plano concatenado), esto lee celda por celda para
# obtener filas estructuradas — necesarias para comparar contra el presupuesto del proyecto.

def _normalizar_col(valor) -> str:
    import unicodedata
    s = str(valor or "").strip().lower()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return re.sub(r"[^a-z0-9]", "", s)


def _parse_precio(valor):
    if valor is None:
        return None
    if isinstance(valor, (int, float)):
        return float(valor)
    s = re.sub(r"[^\d,.\-]", "", str(valor).strip())
    if not s:
        return None
    if "," in s:
        s = s.replace(".", "").replace(",", ".")   # notación chilena: punto=miles, coma=decimal
    elif "." in s:
        s = s.replace(".", "")   # sin coma: el punto es separador de miles, no decimal
    try:
        return float(s)
    except ValueError:
        return None


def parse_tabla_precios(filepath: str) -> list:
    """Lee un Excel de precios referenciales con columnas categoria/item/unidad/precio
    (encabezados case-insensitive, sin tildes, en cualquier orden, en la primera hoja).
    Retorna [{categoria, item, unidad, precio}, ...]. Lanza ValueError si faltan columnas
    obligatorias (categoria, item, precio); unidad es opcional."""
    import openpyxl
    wb = openpyxl.load_workbook(filepath, data_only=True)
    filas = list(wb.worksheets[0].iter_rows(values_only=True))
    if not filas:
        return []

    header = [_normalizar_col(c) for c in filas[0]]
    idx = {}
    for i, h in enumerate(header):
        if h == "categoria" and "categoria" not in idx:
            idx["categoria"] = i
        elif h in ("item", "producto", "descripcion") and "item" not in idx:
            idx["item"] = i
        elif h in ("unidad", "un", "unidadmedida") and "unidad" not in idx:
            idx["unidad"] = i
        elif h in ("precio", "preciopromedio", "valor", "preciounitario") and "precio" not in idx:
            idx["precio"] = i

    faltantes = {"categoria", "item", "precio"} - idx.keys()
    if faltantes:
        raise ValueError(
            "Faltan columnas obligatorias en el Excel: " + ", ".join(sorted(faltantes)) +
            " (encabezados esperados: categoria, item, unidad, precio)")

    items = []
    for row in filas[1:]:
        if row is None:
            continue

        def _get(key):
            i = idx.get(key)
            return row[i] if i is not None and i < len(row) else None

        categoria = str(_get("categoria") or "").strip()
        item = str(_get("item") or "").strip()
        if not categoria or not item:
            continue
        precio = _parse_precio(_get("precio"))
        if precio is None:
            continue
        unidad = str(_get("unidad") or "").strip()
        items.append({"categoria": categoria, "item": item, "unidad": unidad, "precio": precio})
    return items
